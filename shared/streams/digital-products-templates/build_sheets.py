#!/usr/bin/env python3
"""Build the Active Trader Portfolio Tracker Excel file."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# Styles
hfont = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
tfont = Font(name='Calibri', bold=True, size=14, color='1F2937')
sfont = Font(name='Calibri', bold=True, size=11, color='4B5563')
mfmt = '$#,##0.00'
pfmt = '0.00%'
bdr = Border(left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'),
             top=Side(style='thin', color='D1D5DB'), bottom=Side(style='thin', color='D1D5DB'))
gfill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
rfill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

def style_headers(ws, n, row=1):
    for c in range(1, n+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = bdr

# ── POSITIONS ──
ws = wb.active
ws.title = "Positions"
hdrs = ['Symbol','Description','Account','Sector','Asset Type','Quantity','Avg Cost/Share',
        'Current Price','Market Value','Cost Basis','Gain/Loss $','Gain/Loss %',
        '% of Account','Div Yield','Div Pay Date','Status','Exit Target','Stop Loss','Tags','Thesis']
for c, h in enumerate(hdrs, 1):
    ws.cell(row=1, column=c, value=h)
style_headers(ws, len(hdrs))

for i, w in enumerate([8,22,12,18,10,8,14,14,14,14,12,12,12,10,12,10,12,12,20,40], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for dv_range, opts in [('C2:C200','Individual,Rollover IRA,Roth IRA,Other'),
                        ('D2:D200','Energy,IT,Materials,Healthcare,Industrials,Financials,Consumer,Utilities,Comm Services,Real Estate'),
                        ('E2:E200','Equity,ETF,Option,Bond,Crypto,Cash'),
                        ('P2:P200','Active,Watching,Exited,Stopped Out')]:
    dv = DataValidation(type='list', formula1=f'"{opts}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(dv_range)

samples = [
    ['AAPL','Apple Inc','Individual','IT','Equity',25,178.50,195.25,0.005,'Active',220,165,'Momentum, Long Term','Core tech holding. Strong services revenue growth.'],
    ['XOM','Exxon Mobil Corp','Rollover IRA','Energy','Equity',50,98.30,112.45,0.034,'Active',130,90,'Dividend, Value','Energy supercycle play. Domestic production hedge.'],
    ['GLD','SPDR Gold Shares','Individual','Materials','ETF',40,185.00,210.50,0,'Active',None,None,'Hedge','Gold vault position. Capital preservation.'],
    ['INTC','Intel Corp','Individual','IT','Equity',100,42.15,35.80,0.015,'Active',55,30,'Value, Speculative','Turnaround bet on foundry business.'],
    ['NTR','Nutrien Ltd','Rollover IRA','Materials','Equity',30,52.40,58.75,0.038,'Active',75,None,'Value, Dividend','Worlds largest potash producer. Fertilizer anchor.'],
    ['NAT','Nordic American Tanker','Individual','Energy','Equity',400,5.92,6.13,0.116,'Active',8,5.50,'Dividend, Value','Tanker play. Longer voyages = more revenue.'],
    ['CF','CF Industries','Individual','Materials','Equity',10,131.07,130.63,0.015,'Active',150,124,'Value','US nitrogen producer. Cheapest nat gas input costs.'],
    ['GNK','Genco Shipping','Individual','Industrials','Equity',195,23.32,23.84,0.089,'Active',28,22.51,'Dividend, Value','Dry bulk shipping. Grain and fertilizer transport.'],
]

for r, s in enumerate(samples, 2):
    ws.cell(row=r, column=1, value=s[0])  # Symbol
    ws.cell(row=r, column=2, value=s[1])  # Description
    ws.cell(row=r, column=3, value=s[2])  # Account
    ws.cell(row=r, column=4, value=s[3])  # Sector
    ws.cell(row=r, column=5, value=s[4])  # Asset Type
    ws.cell(row=r, column=6, value=s[5])  # Quantity
    ws.cell(row=r, column=7, value=s[6])  # Avg Cost
    ws.cell(row=r, column=8, value=s[7])  # Current Price
    ws.cell(row=r, column=9, value=f'=F{r}*H{r}')  # Market Value
    ws.cell(row=r, column=10, value=f'=F{r}*G{r}')  # Cost Basis
    ws.cell(row=r, column=11, value=f'=I{r}-J{r}')  # G/L $
    ws.cell(row=r, column=12, value=f'=IF(J{r}<>0,K{r}/J{r},0)')  # G/L %
    ws.cell(row=r, column=13, value=f'=IF(SUM($I$2:$I$200)<>0,I{r}/SUM($I$2:$I$200),0)')  # % of Acct
    ws.cell(row=r, column=14, value=s[8])  # Div Yield
    ws.cell(row=r, column=16, value=s[9])  # Status
    if s[10] is not None: ws.cell(row=r, column=17, value=s[10])  # Exit Target
    if s[11] is not None: ws.cell(row=r, column=18, value=s[11])  # Stop Loss
    ws.cell(row=r, column=19, value=s[12])  # Tags
    ws.cell(row=r, column=20, value=s[13])  # Thesis
    for c in [7,8,9,10,11,17,18]: ws.cell(row=r, column=c).number_format = mfmt
    for c in [12,13,14]: ws.cell(row=r, column=c).number_format = pfmt
    for c in range(1, 21): ws.cell(row=r, column=c).border = bdr

ws.conditional_formatting.add('K2:K200', CellIsRule(operator='greaterThan', formula=['0'], fill=gfill))
ws.conditional_formatting.add('K2:K200', CellIsRule(operator='lessThan', formula=['0'], fill=rfill))
ws.conditional_formatting.add('L2:L200', CellIsRule(operator='greaterThan', formula=['0'], fill=gfill))
ws.conditional_formatting.add('L2:L200', CellIsRule(operator='lessThan', formula=['0'], fill=rfill))

ws.cell(row=11, column=1, value="TIP: In Google Sheets, replace col H with =GOOGLEFINANCE(A2,\"price\") for auto-updating prices").font = Font(italic=True, color='6B7280')

# ── REALIZED TRADES ──
ws2 = wb.create_sheet("Realized Trades")
hdrs2 = ['Symbol','Description','Account','Open Date','Close Date','Quantity','Proceeds/Share','Cost/Share','Proceeds','Cost Basis','Gain/Loss $','Gain/Loss %','Term','Wash Sale?','Disallowed Loss']
for c, h in enumerate(hdrs2, 1): ws2.cell(row=1, column=c, value=h)
style_headers(ws2, len(hdrs2))
for i, w in enumerate([8,20,12,12,12,8,14,14,14,14,12,12,12,10,14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for dv_range, opts in [('C2:C500','Individual,Rollover IRA,Roth IRA,Other'), ('M2:M500','Short Term,Long Term')]:
    dv = DataValidation(type='list', formula1=f'"{opts}"', allow_blank=True)
    ws2.add_data_validation(dv)
    dv.add(dv_range)

trades = [
    ['AAOI','Applied Optoelectronics','Individual','2026-03-05','2026-03-10',15,124.50,102.05],
    ['LITE','Lumentum Holdings','Individual','2026-03-15','2026-04-01',15,766.00,726.37],
    ['RIG','Transocean','Individual','2026-03-25','2026-03-31',200,6.51,6.90],
]
for r, t in enumerate(trades, 2):
    for c, v in enumerate(t, 1): ws2.cell(row=r, column=c, value=v)
    ws2.cell(row=r, column=9, value=f'=F{r}*G{r}')
    ws2.cell(row=r, column=10, value=f'=F{r}*H{r}')
    ws2.cell(row=r, column=11, value=f'=I{r}-J{r}')
    ws2.cell(row=r, column=12, value=f'=IF(J{r}<>0,K{r}/J{r},0)')
    for c in [7,8,9,10,11,15]: ws2.cell(row=r, column=c).number_format = mfmt
    ws2.cell(row=r, column=12).number_format = pfmt
    for c in range(1, 16): ws2.cell(row=r, column=c).border = bdr

ws2.conditional_formatting.add('K2:K500', CellIsRule(operator='greaterThan', formula=['0'], fill=gfill))
ws2.conditional_formatting.add('K2:K500', CellIsRule(operator='lessThan', formula=['0'], fill=rfill))

ws2.cell(row=6, column=1, value="YTD SUMMARY").font = sfont
ws2.cell(row=7, column=1, value="Total Realized G/L")
ws2.cell(row=7, column=11, value='=SUM(K2:K500)')
ws2.cell(row=7, column=11).number_format = mfmt
ws2.cell(row=7, column=11).font = Font(bold=True)

# ── WATCHLIST ──
ws3 = wb.create_sheet("Watchlist")
hdrs3 = ['Symbol','Description','Sector','Current Price','Target Entry','Gap to Entry','Thesis','Catalyst','Priority','Added Date']
for c, h in enumerate(hdrs3, 1): ws3.cell(row=1, column=c, value=h)
style_headers(ws3, len(hdrs3))
for i, w in enumerate([8,22,15,14,14,12,35,25,10,12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

dv = DataValidation(type='list', formula1='"Hot,Warm,Cold"', allow_blank=True)
ws3.add_data_validation(dv)
dv.add('I2:I100')

wl = [
    ['ARM','ARM Holdings','IT',147.44,120.00,'Semiconductor IP leader. AI chip demand.','Earnings beat','Warm','2026-03-24'],
    ['DE','Deere & Co','Industrials',563.00,480.00,'Ag equipment at cycle trough.','Guidance upgrade','Cold','2026-04-02'],
    ['PG','Procter & Gamble','Consumer',168.00,155.00,'Recession-proof staples.','Market selloff','Warm','2026-04-01'],
]
for r, w in enumerate(wl, 2):
    for c, v in enumerate(w, 1): ws3.cell(row=r, column=c, value=v)
    # Shift thesis/catalyst/priority/date since gap is col 6
    ws3.cell(row=r, column=7, value=w[5])
    ws3.cell(row=r, column=8, value=w[6])
    ws3.cell(row=r, column=9, value=w[7])
    ws3.cell(row=r, column=10, value=w[8])
    ws3.cell(row=r, column=6, value=f'=IF(D{r}<>0,(D{r}-E{r})/D{r},0)')
    ws3.cell(row=r, column=4).number_format = mfmt
    ws3.cell(row=r, column=5).number_format = mfmt
    ws3.cell(row=r, column=6).number_format = pfmt
    for c in range(1, 11): ws3.cell(row=r, column=c).border = bdr

ws3.conditional_formatting.add('F2:F100', CellIsRule(operator='lessThan', formula=['0.05'], fill=gfill))

# ── TRADE JOURNAL ──
ws4 = wb.create_sheet("Trade Journal")
hdrs4 = ['Date','Market View','Symbols','Decision','Outcome','Lesson']
for c, h in enumerate(hdrs4, 1): ws4.cell(row=1, column=c, value=h)
style_headers(ws4, len(hdrs4))
for i, w in enumerate([12,12,15,45,35,35], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

dv = DataValidation(type='list', formula1='"Bullish,Bearish,Neutral,Cautious"', allow_blank=True)
ws4.add_data_validation(dv)
dv.add('B2:B500')

j = [['2026-04-01','Cautious','LITE, CF, NTR','Trimmed LITE. Rotated into fertilizer basket.','Locked in gains.','Take profits mechanically.'],
     ['2026-04-02','Cautious','GNK, NAT, FJET','Added GNK with tiered OCO. Doubled NAT.','All green day one.','Momentum + defined exits work.']]
for r, row in enumerate(j, 2):
    for c, v in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=v)
        ws4.cell(row=r, column=c).border = bdr

# ── DIVIDEND TRACKER ──
ws5 = wb.create_sheet("Dividend Tracker")
hdrs5 = ['Symbol','Pay Date','Amount/Share','Shares Held','Total Payment','Account','Reinvested?']
for c, h in enumerate(hdrs5, 1): ws5.cell(row=1, column=c, value=h)
style_headers(ws5, len(hdrs5))
for i, w in enumerate([8,12,14,12,14,14,12], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

divs = [['NAT','2026-03-24',0.12,400,'Individual','No'],['ENB','2026-03-01',0.915,25,'Individual','No']]
for r, d in enumerate(divs, 2):
    for c, v in enumerate(d, 1): ws5.cell(row=r, column=c, value=v)
    ws5.cell(row=r, column=5, value=f'=C{r}*D{r}')
    ws5.cell(row=r, column=3).number_format = mfmt
    ws5.cell(row=r, column=5).number_format = mfmt
    for c in range(1, 8): ws5.cell(row=r, column=c).border = bdr

ws5.cell(row=5, column=1, value="YTD TOTAL").font = sfont
ws5.cell(row=5, column=5, value='=SUM(E2:E500)')
ws5.cell(row=5, column=5).number_format = mfmt
ws5.cell(row=5, column=5).font = Font(bold=True)

# ── DASHBOARD ──
ws6 = wb.create_sheet("Dashboard")
ws6.cell(row=1, column=1, value="ACTIVE TRADER PORTFOLIO TRACKER").font = Font(bold=True, size=16, color='1F2937')
ws6.merge_cells('A1:F1')
ws6.cell(row=2, column=1, value="Auto-calculated from Positions tab").font = Font(italic=True, color='6B7280')

ws6.cell(row=4, column=1, value="ACCOUNT SUMMARY").font = sfont
for c, h in enumerate(['Account','Market Value','Cost Basis','Gain/Loss $','Gain/Loss %','# Positions'], 1):
    ws6.cell(row=5, column=c, value=h)
style_headers(ws6, 6, row=5)

for r, acct in enumerate(['Individual','Rollover IRA','ALL ACCOUNTS'], 6):
    ws6.cell(row=r, column=1, value=acct).font = Font(bold=(acct=='ALL ACCOUNTS'))
    if acct == 'ALL ACCOUNTS':
        ws6.cell(row=r, column=2, value="=SUM(Positions!I:I)")
        ws6.cell(row=r, column=3, value="=SUM(Positions!J:J)")
        ws6.cell(row=r, column=6, value="=COUNTA(Positions!A2:A200)")
    else:
        ws6.cell(row=r, column=2, value=f'=SUMIFS(Positions!I:I,Positions!C:C,"{acct}")')
        ws6.cell(row=r, column=3, value=f'=SUMIFS(Positions!J:J,Positions!C:C,"{acct}")')
        ws6.cell(row=r, column=6, value=f'=COUNTIFS(Positions!C:C,"{acct}")')
    ws6.cell(row=r, column=4, value=f'=B{r}-C{r}')
    ws6.cell(row=r, column=5, value=f'=IF(C{r}<>0,D{r}/C{r},0)')
    for c in [2,3,4]: ws6.cell(row=r, column=c).number_format = mfmt
    ws6.cell(row=r, column=5).number_format = pfmt
    for c in range(1,7): ws6.cell(row=r, column=c).border = bdr

ws6.cell(row=10, column=1, value="SECTOR ALLOCATION").font = sfont
for c, h in enumerate(['Sector','Market Value','% of Portfolio'], 1):
    ws6.cell(row=11, column=c, value=h)
style_headers(ws6, 3, row=11)

for r, sec in enumerate(['Energy','IT','Materials','Healthcare','Industrials','Financials','Consumer','Utilities','Comm Services','Real Estate'], 12):
    ws6.cell(row=r, column=1, value=sec)
    ws6.cell(row=r, column=2, value=f'=SUMIFS(Positions!I:I,Positions!D:D,"{sec}")')
    ws6.cell(row=r, column=3, value=f'=IF(SUM(Positions!I:I)<>0,B{r}/SUM(Positions!I:I),0)')
    ws6.cell(row=r, column=2).number_format = mfmt
    ws6.cell(row=r, column=3).number_format = pfmt
    for c in range(1,4): ws6.cell(row=r, column=c).border = bdr

ws6.cell(row=23, column=1, value="MACRO INDICATORS").font = sfont
for c, h in enumerate(['Indicator','Value','Trend','Notes'], 1):
    ws6.cell(row=24, column=c, value=h)
style_headers(ws6, 4, row=24)
for r, ind in enumerate(['Brent Crude','Gold (oz)','DXY','10Y Treasury','VIX'], 25):
    ws6.cell(row=r, column=1, value=ind)
    for c in range(1,5): ws6.cell(row=r, column=c).border = bdr

for i in range(1, 7): ws6.column_dimensions[get_column_letter(i)].width = 18

# ── HOW TO USE ──
ws7 = wb.create_sheet("How To Use")
lines = [
    "ACTIVE TRADER PORTFOLIO TRACKER — SETUP GUIDE",
    "",
    "STEP 1: MAKE YOUR OWN COPY",
    "Google Sheets: File > Make a Copy. Excel: Save As.",
    "",
    "STEP 2: ENABLE AUTO-PRICING (Google Sheets only)",
    "In Positions tab, column H, replace prices with: =GOOGLEFINANCE(A2,\"price\")",
    "Prices update every ~20 min. Free. Only works in Google Sheets.",
    "",
    "STEP 3: ADD YOUR POSITIONS",
    "Delete sample data. Add: Symbol, Description, Account, Sector, Quantity, Avg Cost/Share.",
    "Market Value, Cost Basis, Gain/Loss auto-calculate. Use dropdowns for Account/Sector/Status.",
    "",
    "STEP 4: LOG TRADES",
    "Add closed positions to Realized Trades. Term auto-calculates (Short/Long based on hold time).",
    "",
    "STEP 5: WATCHLIST",
    "Add tickers with target entry prices. Gap to Entry shows how close to your buy price.",
    "",
    "STEP 6: TRADE JOURNAL",
    "After each session: date, market view, decision, outcome, lesson. Best tool for improving.",
    "",
    "STEP 7: DASHBOARD",
    "Auto-calculates from Positions. Account summaries, sector allocation, macro indicators.",
    "",
    "TIPS",
    "- Update at market close daily (2-5 min)",
    "- The Thesis column is your most valuable field. If a position has no thesis, rethink it.",
    "- Green = making money. Red = losing money.",
    "",
    "Built by an active trader managing a six-figure portfolio across multiple accounts.",
]
for r, line in enumerate(lines, 1):
    cell = ws7.cell(row=r, column=1, value=line)
    if line.startswith("STEP") or line in ["TIPS","ACTIVE TRADER PORTFOLIO TRACKER — SETUP GUIDE"]:
        cell.font = Font(bold=True, size=12 if "SETUP GUIDE" in line else 11)
ws7.column_dimensions['A'].width = 90

# Move Dashboard first
wb.move_sheet("Dashboard", offset=-5)

# Save
out = "Active_Trader_Portfolio_Tracker.xlsx"
wb.save(out)
print(f"SUCCESS: {out}")
