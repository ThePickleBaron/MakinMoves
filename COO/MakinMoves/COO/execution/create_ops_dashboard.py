from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
HEADER_FILL = PatternFill('solid', fgColor='1B2A4A')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='2D5AA0')
SECTION_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
LABEL_FONT = Font(name='Arial', size=10)
VALUE_FONT = Font(name='Arial', size=10, color='0000FF')
FORMULA_FONT = Font(name='Arial', size=10, color='000000')
TARGET_FILL = PatternFill('solid', fgColor='E8F0FE')
ALERT_FILL = PatternFill('solid', fgColor='FDECEA')
GREEN_FILL = PatternFill('solid', fgColor='E6F4EA')
YELLOW_FILL = PatternFill('solid', fgColor='FFF8E1')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def style_header(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_range(ws, start_row, end_row, cols):
    for r in range(start_row, end_row+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = LABEL_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True)

# ========== SHEET 1: UNIFIED DASHBOARD ==========
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = '1B2A4A'

headers = ['Metric', 'Today', 'Yesterday', 'Day Δ', 'Week Total', 'Target (Week)', 'vs Target %', 'Status', 'Notes']
col_widths = [30, 12, 12, 10, 14, 14, 12, 10, 25]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title
ws.merge_cells('A1:I1')
ws['A1'] = 'MakinMoves — Unified Operations Dashboard'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1B2A4A')
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:I2')
ws['A2'] = 'Updated: [DATE] | All 3 revenue streams at a glance'
ws['A2'].font = Font(name='Arial', size=9, color='666666')
ws['A2'].alignment = Alignment(horizontal='center')

# P1 GUMROAD Section
row = 4
ws.merge_cells(f'A{row}:I{row}')
ws[f'A{row}'] = '📦 P1 — GUMROAD DIGITAL PRODUCTS'
style_header(ws, row, 9, SECTION_FILL, SECTION_FONT)

row = 5
for i, h in enumerate(headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 9)

p1_metrics = [
    ['Products Live', '', '', '', '', 5, '', '', ''],
    ['Units Sold', '', '', '', '', 3, '', '', ''],
    ['Revenue ($)', '', '', '', '', 75, '', '', 'Net after Gumroad fees'],
    ['Email List Size', '', '', '', '', 50, '', '', ''],
    ['Conversion Rate (%)', '', '', '', '', '3.0%', '', '', 'Visitors → Buyers'],
    ['Customer Reviews', '', '', '', '', 2, '', '', ''],
    ['Discount Uses (LAUNCH25)', '', '', '', '', '-', '', '', '50 total available'],
]

for i, m in enumerate(p1_metrics):
    r = row + 1 + i
    for j, v in enumerate(m):
        ws.cell(row=r, column=j+1, value=v)
    # Day delta formula
    ws.cell(row=r, column=4).value = f'=B{r}-C{r}'
    ws.cell(row=r, column=4).font = FORMULA_FONT
    # vs Target %
    if i < 6:
        ws.cell(row=r, column=7).value = f'=IF(F{r}=0,"-",E{r}/F{r})'
        ws.cell(row=r, column=7).number_format = '0%'
        ws.cell(row=r, column=7).font = FORMULA_FONT
    # Status formula
    ws.cell(row=r, column=8).value = f'=IF(G{r}="-","—",IF(G{r}>=1,"✅",IF(G{r}>=0.7,"⚠️","🔴")))'

style_range(ws, row+1, row+7, 9)

# P2 FREELANCE Section
row = 14
ws.merge_cells(f'A{row}:I{row}')
ws[f'A{row}'] = '✍️ P2 — FREELANCE WRITING SERVICE'
style_header(ws, row, 9, SECTION_FILL, SECTION_FONT)

row = 15
for i, h in enumerate(headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 9)

p2_metrics = [
    ['New Leads', '', '', '', '', 2, '', '', ''],
    ['Qualified Prospects', '', '', '', '', 1, '', '', ''],
    ['Active Clients', '', '', '', '', 1, '', '', 'Target by Day 3'],
    ['Pipeline Value ($)', '', '', '', '', 2400, '', '', 'Total potential revenue'],
    ['MRR ($)', '', '', '', '', 1200, '', '', 'Monthly recurring'],
    ['Invoices Pending ($)', '', '', '', '', '-', '', '', ''],
    ['Invoices Overdue ($)', '', '', '', '', '-', '', '', 'Flag if >$0'],
]

for i, m in enumerate(p2_metrics):
    r = row + 1 + i
    for j, v in enumerate(m):
        ws.cell(row=r, column=j+1, value=v)
    ws.cell(row=r, column=4).value = f'=B{r}-C{r}'
    ws.cell(row=r, column=4).font = FORMULA_FONT
    if i < 5:
        ws.cell(row=r, column=7).value = f'=IF(F{r}=0,"-",E{r}/F{r})'
        ws.cell(row=r, column=7).number_format = '0%'
        ws.cell(row=r, column=7).font = FORMULA_FONT
    ws.cell(row=r, column=8).value = f'=IF(G{r}="-","—",IF(G{r}>=1,"✅",IF(G{r}>=0.7,"⚠️","🔴")))'

style_range(ws, row+1, row+7, 9)

# P3 AFFILIATE Section
row = 24
ws.merge_cells(f'A{row}:I{row}')
ws[f'A{row}'] = '🌐 P3 — AFFILIATE NICHE SITES'
style_header(ws, row, 9, SECTION_FILL, SECTION_FONT)

row = 25
for i, h in enumerate(headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 9)

p3_metrics = [
    ['Articles Published', '', '', '', '', 2, '', '', 'Across all 3 sites'],
    ['Site Traffic (combined)', '', '', '', '', 100, '', '', ''],
    ['Affiliate Clicks', '', '', '', '', 5, '', '', ''],
    ['Affiliate Revenue ($)', '', '', '', '', 10, '', '', ''],
    ['Pages Indexed', '', '', '', '', 10, '', '', 'Google Search Console'],
    ['Backlinks Acquired', '', '', '', '', 2, '', '', ''],
    ['Technical Issues', '', '', '', '', 0, '', '', 'Broken links, 404s, speed'],
]

for i, m in enumerate(p3_metrics):
    r = row + 1 + i
    for j, v in enumerate(m):
        ws.cell(row=r, column=j+1, value=v)
    ws.cell(row=r, column=4).value = f'=B{r}-C{r}'
    ws.cell(row=r, column=4).font = FORMULA_FONT
    if i < 6:
        ws.cell(row=r, column=7).value = f'=IF(F{r}=0,"-",E{r}/F{r})'
        ws.cell(row=r, column=7).number_format = '0%'
        ws.cell(row=r, column=7).font = FORMULA_FONT
    ws.cell(row=r, column=8).value = f'=IF(G{r}="-","—",IF(G{r}>=1,"✅",IF(G{r}>=0.7,"⚠️","🔴")))'

style_range(ws, row+1, row+7, 9)

# OPERATIONAL HEALTH Section
row = 34
ws.merge_cells(f'A{row}:I{row}')
ws[f'A{row}'] = '⚙️ OPERATIONAL HEALTH'
style_header(ws, row, 9, SECTION_FILL, SECTION_FONT)

row = 35
health_headers = ['Check', 'Status', 'Owner', 'Details', 'Last Updated', '', '', '', '']
for i, h in enumerate(health_headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 9)

health_items = [
    ['Critical path on track?', '', 'COO', '', ''],
    ['Resource bottlenecks?', '', 'COO', '', ''],
    ['Team blockers?', '', 'All', '', ''],
    ['Manual workarounds active?', '', 'CTMO', '', ''],
    ['Burn rate vs plan?', '', 'CFO', '', ''],
    ['Unplanned issues today?', '', 'COO', '', ''],
]

for i, item in enumerate(health_items):
    r = row + 1 + i
    for j, v in enumerate(item):
        ws.cell(row=r, column=j+1, value=v)

style_range(ws, row+1, row+6, 5)

# DEVIATION ALERTS Section
row = 43
ws.merge_cells(f'A{row}:I{row}')
ws[f'A{row}'] = '🚨 DEVIATION ALERTS (Auto-flagged when >20% off target)'
style_header(ws, row, 9, PatternFill('solid', fgColor='C62828'), HEADER_FONT)

row = 44
alert_headers = ['Stream', 'Metric', 'Target', 'Actual', 'Deviation %', 'Severity', 'Action Required', '', '']
for i, h in enumerate(alert_headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 9)

for r in range(45, 50):
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.cell(row=r, column=5).value = f'=IF(C{r}=0,"-",(D{r}-C{r})/C{r})'
    ws.cell(row=r, column=5).number_format = '0%'
    ws.cell(row=r, column=6).value = f'=IF(E{r}="-","",IF(ABS(E{r})>0.5,"CRITICAL",IF(ABS(E{r})>0.2,"WARNING","OK")))'

style_range(ws, 45, 49, 9)

# ========== SHEET 2: DAILY UPDATE CHECKLIST ==========
ws2 = wb.create_sheet("Daily Checklist")
ws2.sheet_properties.tabColor = '2D5AA0'
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 30

ws2.merge_cells('A1:E1')
ws2['A1'] = 'Daily Operations Update Checklist (10 min)'
ws2['A1'].font = Font(name='Arial', bold=True, size=13, color='1B2A4A')

daily_items = [
    ('', 'TASK', 'OWNER', 'DONE?', 'NOTES'),
    ('', 'P1 GUMROAD (3 min)', '', '', ''),
    ('1', 'Check Gumroad dashboard — units sold today', 'COO', '☐', ''),
    ('2', 'Update revenue in Dashboard tab', 'COO', '☐', ''),
    ('3', 'Check for customer questions/reviews', 'COO', '☐', ''),
    ('4', 'Update email list count', 'COO', '☐', ''),
    ('', 'P2 FREELANCE (3 min)', '', '', ''),
    ('5', 'Check intake form for new leads', 'COO', '☐', ''),
    ('6', 'Update pipeline stages for active prospects', 'COO', '☐', ''),
    ('7', 'Check invoice payment status', 'CFO', '☐', ''),
    ('8', 'Review any client communications', 'COO', '☐', ''),
    ('', 'P3 AFFILIATE (3 min)', '', '', ''),
    ('9', 'Check Google Analytics — traffic by site', 'COO', '☐', ''),
    ('10', 'Check affiliate dashboards — clicks + revenue', 'COO', '☐', ''),
    ('11', 'Verify scheduled articles published', 'COO', '☐', ''),
    ('12', 'Check Search Console for indexing issues', 'CTMO', '☐', ''),
    ('', 'WRAP-UP (1 min)', '', '', ''),
    ('13', 'Flag any deviation alerts in Dashboard', 'COO', '☐', ''),
    ('14', 'File daily standup to board/standups/', 'COO', '☐', ''),
]

for i, item in enumerate(daily_items):
    r = i + 3
    for j, v in enumerate(item):
        ws2.cell(row=r, column=j+1, value=v)
        ws2.cell(row=r, column=j+1).border = THIN_BORDER
        ws2.cell(row=r, column=j+1).font = LABEL_FONT

    if item[1].startswith('P1') or item[1].startswith('P2') or item[1].startswith('P3') or item[1].startswith('WRAP'):
        for j in range(5):
            ws2.cell(row=r, column=j+1).fill = SECTION_FILL
            ws2.cell(row=r, column=j+1).font = SECTION_FONT

style_header(ws2, 3, 5)

# ========== SHEET 3: WEEKLY REVIEW ==========
ws3 = wb.create_sheet("Weekly Review")
ws3.sheet_properties.tabColor = '0D652D'
ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 15
ws3.column_dimensions['F'].width = 15
ws3.column_dimensions['G'].width = 20

ws3.merge_cells('A1:G1')
ws3['A1'] = 'Weekly Review Template (30 min)'
ws3['A1'].font = Font(name='Arial', bold=True, size=13, color='1B2A4A')

ws3.merge_cells('A2:G2')
ws3['A2'] = 'Week of: [DATE]'
ws3['A2'].font = Font(name='Arial', size=10, color='666666')

row = 4
weekly_headers = ['Metric', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Week Total']
for i, h in enumerate(weekly_headers, 1):
    ws3.cell(row=row, column=i, value=h)
style_header(ws3, row, 7)

weekly_metrics = [
    'P1: Units Sold', 'P1: Revenue ($)', 'P1: New Email Subs',
    'P2: New Leads', 'P2: Clients Closed', 'P2: Revenue ($)',
    'P3: Articles Published', 'P3: Site Traffic', 'P3: Affiliate Revenue ($)',
    'TOTAL Revenue ($)'
]

for i, metric in enumerate(weekly_metrics):
    r = row + 1 + i
    ws3.cell(row=r, column=1, value=metric)
    ws3.cell(row=r, column=7).value = f'=SUM(B{r}:F{r})'
    ws3.cell(row=r, column=7).font = FORMULA_FONT
    for c in range(1, 8):
        ws3.cell(row=r, column=c).border = THIN_BORDER
        ws3.cell(row=r, column=c).font = LABEL_FONT

# Trend section
trend_row = row + len(weekly_metrics) + 2
ws3.merge_cells(f'A{trend_row}:G{trend_row}')
ws3[f'A{trend_row}'] = 'WEEK-OVER-WEEK TRENDS'
style_header(ws3, trend_row, 7, SECTION_FILL, SECTION_FONT)

trend_headers = ['Metric', 'Last Week', 'This Week', 'Change', 'Change %', 'Trend', 'Action Needed?']
tr = trend_row + 1
for i, h in enumerate(trend_headers, 1):
    ws3.cell(row=tr, column=i, value=h)
style_header(ws3, tr, 7)

trend_metrics = ['Total Revenue', 'Total Units/Clients', 'Total Traffic', 'Email List Size', 'Conversion Rate']
for i, m in enumerate(trend_metrics):
    r = tr + 1 + i
    ws3.cell(row=r, column=1, value=m)
    ws3.cell(row=r, column=4).value = f'=C{r}-B{r}'
    ws3.cell(row=r, column=4).font = FORMULA_FONT
    ws3.cell(row=r, column=5).value = f'=IF(B{r}=0,"-",D{r}/B{r})'
    ws3.cell(row=r, column=5).number_format = '0%'
    ws3.cell(row=r, column=5).font = FORMULA_FONT
    ws3.cell(row=r, column=6).value = f'=IF(D{r}>0,"📈",IF(D{r}<0,"📉","➡️"))'
    for c in range(1, 8):
        ws3.cell(row=r, column=c).border = THIN_BORDER
        ws3.cell(row=r, column=c).font = LABEL_FONT

# Decisions section
dec_row = tr + len(trend_metrics) + 2
ws3.merge_cells(f'A{dec_row}:G{dec_row}')
ws3[f'A{dec_row}'] = 'DECISIONS TO MAKE THIS WEEK'
style_header(ws3, dec_row, 7, PatternFill('solid', fgColor='E65100'), HEADER_FONT)

dec_headers = ['Decision', 'Owner', 'Deadline', 'Options', 'Chosen', 'Rationale', 'Status']
dr = dec_row + 1
for i, h in enumerate(dec_headers, 1):
    ws3.cell(row=dr, column=i, value=h)
style_header(ws3, dr, 7)

for r in range(dr+1, dr+6):
    for c in range(1, 8):
        ws3.cell(row=r, column=c).border = THIN_BORDER

# ========== SHEET 4: ALERT THRESHOLDS ==========
ws4 = wb.create_sheet("Alert Thresholds")
ws4.sheet_properties.tabColor = 'C62828'
ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 15
ws4.column_dimensions['F'].width = 30

ws4.merge_cells('A1:F1')
ws4['A1'] = 'Deviation Alert Thresholds'
ws4['A1'].font = Font(name='Arial', bold=True, size=13, color='C62828')

row = 3
thresh_headers = ['Stream', 'Metric', 'Warning (>20% off)', 'Critical (>50% off)', 'Escalate To', 'Action']
for i, h in enumerate(thresh_headers, 1):
    ws4.cell(row=row, column=i, value=h)
style_header(ws4, row, 6)

thresholds = [
    ['P1', 'Daily Revenue', '<$10/day by Day 3', '<$5/day by Day 5', 'CEO', 'Adjust pricing or messaging'],
    ['P1', 'Conversion Rate', '<2% after 100 visitors', '<1% after 200 visitors', 'COO+CEO', 'A/B test product page'],
    ['P1', 'Email List Growth', '<10 subs/day by Day 3', '<5 subs/day by Day 5', 'COO', 'Add lead magnet'],
    ['P2', 'Lead Generation', '<1 lead/day by Day 5', '0 leads by Day 7', 'CEO', 'Pivot outreach strategy'],
    ['P2', 'Client Close Rate', '<25% of qualified leads', '<10% of qualified leads', 'CEO', 'Review pricing/offer'],
    ['P2', 'Invoice Collection', 'Any invoice >7 days late', 'Any invoice >14 days late', 'CFO', 'Payment follow-up sequence'],
    ['P3', 'Content Velocity', '<2 articles/week', '<1 article/week', 'COO', 'Hire writer or simplify'],
    ['P3', 'Traffic Growth', '<50 visits/day by Week 2', '<20 visits/day by Week 3', 'CTMO', 'SEO audit + content refresh'],
    ['P3', 'Indexation', '<50% pages indexed by Week 2', '<25% by Week 3', 'CTMO', 'Technical SEO fix'],
    ['ALL', 'Total Weekly Revenue', '<$100/week by Week 2', '<$50/week by Week 3', 'CEO', 'Strategic review of all streams'],
    ['ALL', 'Burn Rate', 'Spending >$50/week on tools', 'Spending >$100/week', 'CFO', 'Cost audit'],
]

for i, t in enumerate(thresholds):
    r = row + 1 + i
    for j, v in enumerate(t):
        ws4.cell(row=r, column=j+1, value=v)
        ws4.cell(row=r, column=j+1).border = THIN_BORDER
        ws4.cell(row=r, column=j+1).font = LABEL_FONT

output = '/sessions/nifty-adoring-turing/mnt/MakinMoves/COO/execution/unified_ops_dashboard_2026-03-29.xlsx'
wb.save(output)
print(f"Saved to {output}")
